from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MIN_COLUMN_WIDTH = 15
MAX_EXPECTED_ROWS = 100


def read_sheet(ws):
    table = []
    header = None
    for row in ws.values:
        if not header:
            header = list(row)
            continue
        newrow = OrderedDict()
        for i in range(0, min(len(header), len(row))):
            newrow[header[i]] = row[i] or ""
        table.append(newrow)
    return table


def read(path, sheet=None):
    wb = load_workbook(path)
    ws = wb.active
    if sheet:
        ws = wb[sheet]
    return read_sheet(ws)


def write(grids, output):
    """Given a list of grids and a file-like output, save an XLSX file.
    In addition to "headers" and "rows", the grid may contain these keys:
    - title string: sets the sheet title in Excel
    - active bool: sets the active sheet in Excel
    - activeCell string: sets the selected cell in Excel
    - locked bool: locks the sheet in Excel"""
    bold = Font(bold=True)
    # yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # orange = PatternFill(start_color="FFF1D8", end_color="FFF1D8", fill_type="solid")
    red = PatternFill(start_color="FFD8D8", end_color="FFD8D8", fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)

    for grid in grids:
        ws = None
        if "title" in grid:
            ws = wb.create_sheet(grid["title"])
        else:
            ws = wb.create_sheet()
        grid["worksheet"] = ws
        i = 0
        if "headers" in grid:
            for header in grid["headers"]:
                for j in range(0, len(header)):
                    value = header[j]["label"]
                    cell = ws.cell(column=j + 1, row=i + 1)
                    cell.value = value
                    cell.font = bold
                    c = get_column_letter(j + 1)
                    if i == 0:
                        w = max(len(value), MIN_COLUMN_WIDTH)
                        ws.column_dimensions[c].width = w
                    if "locked" in header[j] and header[j]["locked"]:
                        cell.protection = Protection(locked=True)
                    if "validations" in header[j]:
                        x = len(grid["headers"]) + 1
                        for validation in header[j]["validations"]:
                            data_val = DataValidation(**validation)
                            ws.add_data_validation(data_val)
                            data_val.add("{0}{1}:{0}{2}".format(c, x, MAX_EXPECTED_ROWS))
                i += 1
            ws.freeze_panes = ws.cell(column=1, row=i + 1)

        if "rows" in grid:
            for row in grid["rows"]:
                for j in range(0, len(row)):
                    c = row[j]
                    cell = ws.cell(column=j + 1, row=i + 1)
                    cell.value = c["label"]
                    if "status" in c and c["status"] == "ERROR":
                        cell.fill = red
                    if "comment" in c:
                        cell.comment = Comment(c["comment"], "Validation service")
                    if "bold" in c and c["bold"]:
                        cell.font = bold
                    if "width" in c and c["width"]:
                        a = get_column_letter(j + 1)
                        ws.column_dimensions[a].width = c["width"]
                i += 1

    for grid in grids:
        if "active" in grid and grid["active"]:
            wb.active = grid["worksheet"]
        if "activeCell" in grid:
            grid["worksheet"].sheet_view.selection[0].activeCell = grid["activeCell"]
            grid["worksheet"].sheet_view.selection[0].sqref = grid["activeCell"]
        if "locked" in grid and grid["locked"]:
            grid["worksheet"].protection.enable()

    wb.save(output)
