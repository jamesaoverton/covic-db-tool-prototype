import tables

from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.comments import Comment

MIN_COLUMN_WIDTH = 15


def read_sheet(ws):
    table = []
    header = None
    for row in ws.values:
        if not header:
            header = list(row)
            continue
        newrow = OrderedDict()
        for i in range(0, min(len(header), len(row))):
            newrow[header[i]] = row[i] or ""
        table.append(newrow)
    return table


def read_xlsx(path, sheet=None):
    wb = load_workbook(path)
    ws = wb.active
    if sheet:
        ws = wb[sheet]
    return read_sheet(ws)


def write_xlsx(sheets, path):
    """Given a list of sheets (pairs of title string and grid) and a path,
    write an XLSX file."""
    bold = Font(bold=True)
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange = PatternFill(start_color="FFF1D8", end_color="FFF1D8", fill_type="solid")
    red = PatternFill(start_color="FFD8D8", end_color="FFD8D8", fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)

    for title, grid in sheets:
        ws = wb.create_sheet(title)
        i = 0
        if "headers" in grid:
            for header in grid["headers"]:
                for j in range(0, len(header)):
                    value = header[j]["label"]
                    cell = ws.cell(column=j + 1, row=i + 1)
                    cell.value = value
                    cell.font = bold
                    if i == 0:
                        c = get_column_letter(j + 1)
                        w = max(len(value), MIN_COLUMN_WIDTH)
                        ws.column_dimensions[c].width = w
                i += 1
        ws.freeze_panes = ws.cell(column=1, row=i + 1)
        if "rows" in grid:
            for row in grid["rows"]:
                print("ROW", row)
                for j in range(0, len(row)):
                    c = row[j]
                    cell = ws.cell(column=j + 1, row=i + 1)
                    cell.value = c["label"]
                    if "status" in c and c["status"] == "ERROR":
                        cell.fill = red
                    if "comment" in c:
                        cell.comment = Comment(c["comment"], "Validation service")
                i += 1

    wb.save(filename=path)
